#!/usr/bin/env python3
"""Build a gap list of drug names missing from FAISS metadata."""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_FAISS_META_PATH = ROOT / "data" / "faiss_meta.json"
DEFAULT_EXCEL_PATH = (
    ROOT / "경구약제 이미지 데이터(데이터 설명서, 경구약제 리스트)" / "단일 경구약제_5,000종 리스트.xlsx"
)
DEFAULT_OUT_CSV = ROOT / "runs" / "reports" / "drug_reference_gap_list.csv"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build drug reference gap CSV from FAISS meta + excel list")
    parser.add_argument("--faiss-meta-path", type=Path, default=DEFAULT_FAISS_META_PATH)
    parser.add_argument("--excel-path", type=Path, default=DEFAULT_EXCEL_PATH)
    parser.add_argument("--out-csv", type=Path, default=DEFAULT_OUT_CSV)
    return parser.parse_args()


def normalize_name_for_match(value: str) -> str:
    text = (value or "").strip().lower()
    if not text:
        return ""
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(
        r"\d+(?:[/-]\d+)+(?:\.\d+)?\s*(mg|ml|g|mcg|μg|iu)\b",
        " ",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(r"\b\d+(?:\.\d+)?\s*(mg|ml|g|mcg|μg|iu)\b", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"(?<=[가-힣a-z0-9])[/-]+(?=\s|$)", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\b\d+\s*[tc]\b", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip()
    return text.replace(" ", "")


def load_faiss_normalized_names(path: Path) -> set[str]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    entries = payload.get("entries", [])
    if not isinstance(entries, list):
        return set()

    normalized: set[str] = set()
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        name = str(entry.get("name", "")).strip()
        key = normalize_name_for_match(name)
        if key:
            normalized.add(key)
    return normalized


def read_excel_rows(excel_path: Path) -> tuple[int, list[dict[str, str]]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required. install with: pip install openpyxl") from exc

    wb = openpyxl.load_workbook(str(excel_path), read_only=True)
    rows: list[dict[str, str]] = []
    seen_codes: set[str] = set()
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            total_rows += 1
            c_code = str(row[0]).replace("'", "").strip() if row and row[0] else ""
            name = str(row[2]).replace("'", "").strip() if row and row[2] else ""
            if not c_code or not name or c_code in seen_codes:
                continue
            seen_codes.add(c_code)
            rows.append({"c_code": c_code, "name": name})

    wb.close()
    return total_rows, rows


def build_gap_rows(candidates: list[dict[str, str]], faiss_names: set[str]) -> list[dict[str, str]]:
    gap_rows: list[dict[str, str]] = []
    for row in candidates:
        c_code = row["c_code"]
        name = row["name"]
        normalized_name = normalize_name_for_match(name)
        if not normalized_name:
            continue
        if normalized_name in faiss_names:
            continue
        gap_rows.append(
            {
                "c_code": c_code,
                "name": name,
                "normalized_name": normalized_name,
                "in_faiss": "false",
                "reason": "name_not_in_faiss",
            }
        )
    return gap_rows


def write_gap_csv(rows: list[dict[str, str]], out_csv: Path) -> None:
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["c_code", "name", "normalized_name", "in_faiss", "reason"]
    with out_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    args = parse_args()

    total_excel_rows, candidate_rows = read_excel_rows(args.excel_path)
    faiss_names = load_faiss_normalized_names(args.faiss_meta_path)
    gap_rows = build_gap_rows(candidate_rows, faiss_names)
    write_gap_csv(gap_rows, args.out_csv)

    gap_ratio = (len(gap_rows) / len(candidate_rows) * 100.0) if candidate_rows else 0.0
    sample_names = [row["name"] for row in gap_rows[:10]]

    print(f"[STAT] total_excel_rows={total_excel_rows}")
    print(f"[STAT] faiss_name_count={len(faiss_names)}")
    print(f"[STAT] candidate_rows_after_dedup={len(candidate_rows)}")
    print(f"[STAT] gap_rows={len(gap_rows)}")
    print(f"[STAT] gap_ratio_percent={gap_ratio:.2f}")
    print(f"[STAT] sample_gap_names={json.dumps(sample_names, ensure_ascii=False)}")
    print(f"[OUT] csv={args.out_csv}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
