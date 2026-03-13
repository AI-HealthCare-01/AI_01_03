#!/usr/bin/env python3
"""
의약품 공공데이터(e약은요) -> drug_references 초기 적재 스크립트.

사용 예시:
  PYTHONPATH=. python scripts/fetch_drug_reference_to_db.py --limit 100 --db-host 127.0.0.1
  PYTHONPATH=. python scripts/fetch_drug_reference_to_db.py --limit 20 --dry-run
"""

from __future__ import annotations

import argparse
import asyncio
import json
import re
import sys
import time
import urllib.parse
import urllib.request
from copy import deepcopy
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from tortoise import Tortoise

from app.db.databases import TORTOISE_ORM
from app.models.drug_reference import DrugReference

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_EXCEL_PATH = ROOT / "경구약제 이미지 데이터(데이터 설명서, 경구약제 리스트)" / "단일 경구약제_5,000종 리스트.xlsx"
API_BASE = "https://apis.data.go.kr/1471000/DrbEasyDrugInfoService/getDrbEasyDrugList"

FIELD_MAP: dict[str, str] = {
    "efcyQesitm": "efficacy_text",
    "useMethodQesitm": "dosage_text",
    "atpnQesitm": "precautions_text",
    "atpnWarnQesitm": "warnings_text",
    "intrcQesitm": "interactions_text",
    "seQesitm": "side_effects_text",
    "depositMethodQesitm": "storage_text",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fetch e약은요 data and upsert into drug_references")
    parser.add_argument("--excel-path", type=Path, default=DEFAULT_EXCEL_PATH)
    parser.add_argument("--limit", type=int, default=0, help="0 means all")
    parser.add_argument("--sleep", type=float, default=0.3, help="sleep seconds between API calls")
    parser.add_argument("--dry-run", action="store_true", help="fetch/map only, do not write DB")
    parser.add_argument("--start-index", type=int, default=0)
    parser.add_argument("--source", type=str, default="easy_drug")
    parser.add_argument("--db-host", type=str, default=None, help="optional DB host override")
    return parser.parse_args()


def load_env() -> dict[str, str]:
    env: dict[str, str] = {}
    env_path = ROOT / ".env"
    if not env_path.exists():
        return env
    for line in env_path.read_text(encoding="utf-8").splitlines():
        text = line.strip()
        if not text or text.startswith("#") or "=" not in text:
            continue
        key, value = text.split("=", 1)
        env[key.strip()] = value.strip()
    return env


def strip_html(text: str | None) -> str:
    if not text:
        return ""
    value = re.sub(r"<[^>]+>", "", text)
    value = value.replace("\xa0", " ")
    value = re.sub(r"\n{3,}", "\n\n", value)
    return value.strip()


def normalize_medication_id(c_code: str, name: str) -> str:
    code = re.sub(r"[^A-Z0-9]+", "_", (c_code or "").upper()).strip("_")
    code = re.sub(r"_+", "_", code)
    if code:
        return code

    fallback = re.sub(r"[^A-Z0-9]+", "_", (name or "").upper()).strip("_")
    fallback = re.sub(r"_+", "_", fallback)
    return fallback or "UNKNOWN_PILL"


def build_search_names(product_name: str) -> list[str]:
    candidates: list[str] = []
    seen: set[str] = set()
    name = product_name.strip()

    def add(value: str) -> None:
        s = value.strip()
        if s and s not in seen and len(s) >= 2:
            candidates.append(s)
            seen.add(s)

    no_paren = re.sub(r"\s*\([^)]*\)", "", name).strip()
    no_qty = re.sub(r"\s+\d+[TCc정캡].*$", "", no_paren).strip()
    no_dose = re.sub(r"\s*[\d./-]+\s*(mg|MG|밀리그.*|IU|mcg|MCG|μg|mL|ML)\s*$", "", no_qty).strip()
    no_dose = re.sub(r"\s+[\d./-]+\s*$", "", no_dose).strip()

    add(no_paren)
    add(no_qty)
    add(no_dose)

    for base in (no_qty, no_dose):
        if len(base) > 4:
            match = re.match(r"^[가-힣]{2,4}(?=.{3,})", base)
            if match:
                stripped = base[match.end() :]
                add(stripped)
                stripped_no_dose = re.sub(
                    r"\s*[\d./-]+\s*(mg|MG|밀리그.*|IU|mcg|MCG|μg|mL|ML)\s*$", "", stripped
                ).strip()
                stripped_no_dose = re.sub(r"\s+[\d./-]+\s*$", "", stripped_no_dose).strip()
                add(stripped_no_dose)

    return candidates


def fetch_drug_detail(api_key: str, product_name: str) -> dict[str, Any] | None:
    for search_name in build_search_names(product_name):
        try:
            params = {
                "serviceKey": api_key,
                "itemName": search_name,
                "type": "json",
                "numOfRows": "3",
            }
            url = f"{API_BASE}?{urllib.parse.urlencode(params, safe='%+')}"
            req = urllib.request.Request(url)
            with urllib.request.urlopen(req, timeout=10) as response:
                data = json.loads(response.read().decode("utf-8"))

            items = data.get("body", {}).get("items", [])
            if not items:
                continue

            item = items[0]
            result: dict[str, Any] = {
                "drug_name": item.get("itemName", "") or product_name,
                "company_name": item.get("entpName", ""),
                "source_item_seq": str(item.get("itemSeq", "") or ""),
                "raw_payload_json": item,
            }
            for api_field, model_field in FIELD_MAP.items():
                result[model_field] = strip_html(item.get(api_field))
            return result
        except Exception:
            continue

    return None


def read_excel(excel_path: Path) -> list[dict[str, str]]:
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] openpyxl is required. install with: pip install openpyxl")
        raise

    wb = openpyxl.load_workbook(str(excel_path), read_only=True)
    rows: list[dict[str, str]] = []
    seen_codes: set[str] = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            c_code = str(row[0]).replace("'", "").strip() if row and row[0] else ""
            category = str(row[1]).replace("'", "").strip() if row and row[1] else ""
            name = str(row[2]).replace("'", "").strip() if row and row[2] else ""

            if not c_code or not name or c_code in seen_codes:
                continue

            seen_codes.add(c_code)
            rows.append({"c_code": c_code, "category": category, "name": name})

    wb.close()
    return rows


def build_tortoise_config(db_host: str | None) -> dict[str, Any]:
    cfg = deepcopy(TORTOISE_ORM)
    if db_host:
        cfg["connections"]["default"]["credentials"]["host"] = db_host
    return cfg


async def upsert_payloads(payloads: list[dict[str, Any]], db_host: str | None) -> tuple[int, int, int]:
    inserted = 0
    updated = 0
    errors = 0

    await Tortoise.init(config=build_tortoise_config(db_host))
    try:
        for payload in payloads:
            try:
                medication_id = payload["medication_id"]
                existing = await DrugReference.get_or_none(medication_id=medication_id)
                if existing is None:
                    await DrugReference.create(**payload)
                    inserted += 1
                else:
                    for key, value in payload.items():
                        setattr(existing, key, value)
                    await existing.save()
                    updated += 1
            except Exception:
                errors += 1
    finally:
        await Tortoise.close_connections()

    return inserted, updated, errors


def main() -> int:
    args = parse_args()
    started_at = time.perf_counter()

    stats: dict[str, int] = {
        "total_rows": 0,
        "target_rows": 0,
        "processed": 0,
        "api_success": 0,
        "api_not_found_skipped": 0,
        "row_invalid_skipped": 0,
        "upsert_inserted": 0,
        "upsert_updated": 0,
        "errors": 0,
    }

    try:
        env = load_env()
        api_key = env.get("DATA_GO_KR_API_KEY_ENCODED", "")
        if not api_key:
            print("[ERROR] DATA_GO_KR_API_KEY_ENCODED is missing in .env")
            return 1

        all_rows = read_excel(args.excel_path)
        stats["total_rows"] = len(all_rows)

        sliced = all_rows[max(0, args.start_index) :]
        if args.limit and args.limit > 0:
            sliced = sliced[: args.limit]
        stats["target_rows"] = len(sliced)

        payloads: list[dict[str, Any]] = []

        for row in sliced:
            stats["processed"] += 1
            c_code = (row.get("c_code") or "").strip()
            name = (row.get("name") or "").strip()
            if not c_code or not name:
                stats["row_invalid_skipped"] += 1
                continue

            detail = fetch_drug_detail(api_key=api_key, product_name=name)
            if not detail:
                stats["api_not_found_skipped"] += 1
                if args.sleep > 0:
                    time.sleep(args.sleep)
                continue

            stats["api_success"] += 1
            medication_id = normalize_medication_id(c_code=c_code, name=name)
            payloads.append(
                {
                    "medication_id": medication_id,
                    "drug_name": (detail.get("drug_name") or name)[:255],
                    "company_name": (detail.get("company_name") or "")[:255] or None,
                    "efficacy_text": detail.get("efficacy_text") or None,
                    "dosage_text": detail.get("dosage_text") or None,
                    "precautions_text": detail.get("precautions_text") or None,
                    "warnings_text": detail.get("warnings_text") or None,
                    "interactions_text": detail.get("interactions_text") or None,
                    "side_effects_text": detail.get("side_effects_text") or None,
                    "storage_text": detail.get("storage_text") or None,
                    "source": (args.source or "easy_drug")[:50],
                    "source_item_seq": (detail.get("source_item_seq") or "")[:100] or None,
                    "raw_payload_json": detail.get("raw_payload_json") or {},
                    "last_synced_at": datetime.now(UTC),
                }
            )

            if args.sleep > 0:
                time.sleep(args.sleep)

        if not args.dry_run:
            inserted, updated, upsert_errors = asyncio.run(upsert_payloads(payloads, args.db_host))
            stats["upsert_inserted"] = inserted
            stats["upsert_updated"] = updated
            stats["errors"] += upsert_errors

        elapsed_seconds = time.perf_counter() - started_at
        print(f"[STAT] total_rows={stats['total_rows']}")
        print(f"[STAT] target_rows={stats['target_rows']}")
        print(f"[STAT] processed={stats['processed']}")
        print(f"[STAT] api_success={stats['api_success']}")
        print(f"[STAT] api_not_found_skipped={stats['api_not_found_skipped']}")
        print(f"[STAT] row_invalid_skipped={stats['row_invalid_skipped']}")
        print(f"[STAT] upsert_inserted={stats['upsert_inserted']}")
        print(f"[STAT] upsert_updated={stats['upsert_updated']}")
        print(f"[STAT] errors={stats['errors']}")
        print(f"[STAT] elapsed_seconds={elapsed_seconds:.2f}")
        print(f"[STAT] dry_run={args.dry_run}")

        return 1 if stats["errors"] > 0 else 0
    except Exception as exc:
        print(f"[ERROR] {type(exc).__name__}: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
